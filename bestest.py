import json
import boto3
import datetime
import openpyxl

def convert_excel_to_json(sheet):
    # Get the column names from the first row
    column_names = [cell.value for cell in sheet[2]]

    # Initialize an empty list to store the JSON objects
    json_data = []

    # Iterate over the rows starting from the third row
    for row in sheet.iter_rows(min_row=3, values_only=True):
        row_data = {}
        for col_index, cell_value in enumerate(row):
            # Evaluate formula values if applicable
            if isinstance(cell_value, str) and cell_value.startswith('='):
                cell = sheet.cell(row=col_index + 3, column=col_index + 2)
                cell_value = cell.value

            # Convert date values to the desired format
            if isinstance(cell_value, datetime.datetime):
                cell_value = cell_value.strftime('%d-%b-%y')

            row_data[column_names[col_index]] = cell_value

        json_data.append(row_data)

    return json_data

def upload_json_to_s3(json_data, bucket_name, s3_key):
    # Create an S3 client
    s3_client = boto3.client('s3')

    # Convert the data to JSON
    json_string = json.dumps(json_data)

    # Upload the JSON data to the S3 bucket
    s3_client.put_object(Body=json_string, Bucket=bucket_name, Key=s3_key)

# Provide the path to your Excel file
excel_file_path = "./Thike.xlsx.xlsm"

# Load the Excel workbook
workbook = openpyxl.load_workbook(excel_file_path, data_only=True)

# Get the current date
current_date = datetime.datetime.now().strftime('%Y-%m-%d')

# Iterate over the sheets in the workbook
for sheet in workbook:
    # Convert Excel data to JSON for each sheet
    json_data = convert_excel_to_json(sheet)

    # Create the S3 key using the current date and sheet name
    s3_key = f"{current_date}_{sheet.title}.json"

    # Provide your S3 bucket name
    bucket_name = "bucket-for-macro"

    # Upload the JSON file to S3 for each sheet
    upload_json_to_s3(json_data, bucket_name, s3_key)
